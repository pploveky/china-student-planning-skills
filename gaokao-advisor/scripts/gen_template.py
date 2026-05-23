#!/usr/bin/env python3
"""高考志愿批量录入模板生成脚本。

用途：
- 生成 students-template.xlsx 空模板，供批量录入高考志愿咨询学生信息。

核心逻辑：
1. 复用 gen_test_dataset.py 中的 29 列 HEADERS，保证模板列顺序与解析校验脚本一致。
2. 写入 1 行浅黄色示例数据，提示用户复制模板后删除示例行。
3. 生成“字段说明”sheet，列出字段类型、是否必填、取值范围和填写说明。

输入输出：
- 输入：可选输出文件路径；不传时默认生成 students-template.xlsx。
- 输出：包含 students sheet 与字段说明 sheet 的 xlsx 文件。

关键依赖：
- Python 3.10+
- openpyxl

最小使用示例：
    python3 gen_template.py assets/students-template.xlsx
"""
from __future__ import annotations
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 复用 gen_test_dataset 的 HEADERS 与 field_doc
sys.path.insert(0, str(Path(__file__).parent))
from gen_test_dataset import HEADERS  # noqa: E402


EXAMPLE_ROW = [
    "张同学", "示例-01", "学生本人",
    "北京", "北京", 2026, "物化生",
    670, 1500, 125, 140, 142, "英语",
    "计算机/AI/通信", "", "清华大学",
    "内向独立", "北京/长三角", "",
    "学校>专业>城市",
    6.0, "否", "是", "是",
    "国内读研", "5+", "无", "普通批",
    "示例数据，请删除后填写真实学生信息",
]


def write_template(out_path: Path) -> None:
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "students"
    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    # 1 行示例（用浅黄色背景标识）
    ws.append(EXAMPLE_ROW)
    example_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in ws[2]:
        cell.fill = example_fill

    widths = {
        "A": 22, "B": 9, "C": 11, "D": 8, "E": 8, "F": 10, "G": 12,
        "H": 7, "I": 10, "J": 7, "K": 7, "L": 7, "M": 9,
        "N": 18, "O": 16, "P": 12, "Q": 11, "R": 18, "S": 12, "T": 16,
        "U": 13, "V": 14, "W": 12, "X": 12, "Y": 11, "Z": 12,
        "AA": 14, "AB": 16, "AC": 32,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Sheet 2: 字段说明（直接复用 gen_test_dataset 的 field_doc 块）
    from gen_test_dataset import write_xlsx as _wx  # noqa: F401  仅为复用 doc 元组
    # 直接 inline field_doc，避免循环 import
    field_doc = [
        ("学生姓名", "文本", "是", "—", "用作输出 PDF 文件名"),
        ("学生代号", "文本", "否", "—", "隐私脱敏；留空则用真名"),
        ("主要决策者", "枚举", "是", "学生本人/家长/共同", "影响报告语气"),
        ("省份", "文本", "是", "31 省+港澳台", "决定一分一段表数据源"),
        ("户籍", "文本", "是", "同省份", "决定专项资格"),
        ("高考年份", "整数", "是", "2024-2030", "决定数据年份"),
        ("科类_选科", "文本", "是",
         "老高考: 文科/理科\n新高考: 物化生 物化地 物生地 史政地 史地生 等", "—"),
        ("总分", "整数", "是", "0-900", "海南 900 制；其它省 750 制"),
        ("全省位次", "整数", "是", "1-1500000；0 表未知", "0 由主流程反推"),
        ("语文分", "整数", "是", "0-150", "单科隐性门槛判定用"),
        ("数学分", "整数", "是", "0-150", "同上"),
        ("外语分", "整数", "是", "0-150", "外语类专业普遍 ≥120/130"),
        ("外语语种", "枚举", "否", "英语/日语/俄语/德语/法语/西语", "缺省英语"),
        ("兴趣方向", "文本", "是", "自由文本", "如 计算机/AI/金融"),
        ("厌恶清单", "文本", "否", "自由文本，逗号分隔", "如 生化环材, 师范"),
        ("梦想校", "文本", "否", "—", "用于诚实评估+替代路径"),
        ("性格倾向", "枚举", "是", "内向独立/内向协作/外向独立/外向协作", "—"),
        ("地域偏好", "文本", "是", "自由文本", "如 北京/长三角/不限"),
        ("排斥地域", "文本", "否", "自由文本", "如 东北,西北"),
        ("优先级排序", "文本", "是", "用 > 分隔的 3 项",
         "如 学校>专业>城市"),
        ("学费上限_万每年", "数字", "是", "0 表无限制", "—"),
        ("接受民办_独立学院", "布尔", "是", "是/否", "—"),
        ("接受中外合办", "布尔", "是", "是/否", "—"),
        ("服从专业调剂", "布尔", "是", "是/否", "—"),
        ("读研意向", "枚举", "是", "直接就业/国内读研/出国/不确定", "—"),
        ("家庭月收入_万", "文本", "是", "区间 如 1-2 / 2-5 / 5+", "—"),
        ("体检受限项", "文本", "是", "无 或 色弱;视力<4.8（分号分隔）", "—"),
        ("特殊招生通道", "文本", "是", "普通批 或 强基:数学;综评:港中深", "—"),
        ("备注", "文本", "否", "自由文本", "供主流程吸收"),
    ]

    ws2 = wb.create_sheet("字段说明")
    ws2.append(["字段名", "类型", "必填", "枚举/取值", "说明"])
    for c in ws2[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
    for row in field_doc:
        ws2.append(list(row))
    for col, w in {"A": 18, "B": 8, "C": 6, "D": 32, "E": 36}.items():
        ws2.column_dimensions[col].width = w
    for r in ws2.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(out_path)
    print(f"已生成 {out_path}")


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("students-template.xlsx")
    write_template(out)
