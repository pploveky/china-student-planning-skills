#!/usr/bin/env python3
"""高考志愿批量学生信息解析与校验脚本。

用途：
- 读取批量咨询 xlsx 中的学生画像，输出可供高考志愿规划流程逐行处理的结构化数据。

核心逻辑：
1. 读取首个 sheet 中的 29 列学生信息，其它说明 sheet 自动忽略。
2. 按 SCHEMA 执行字段类型转换、枚举校验、必填校验和数值范围校验。
3. 合法行进入 rows，错误行进入 errors；单行失败不阻断其它学生。

输入输出：
- 输入：students-template.xlsx 或同结构 xlsx 文件。
- 输出：rows 学生字典列表、errors 行错误清单；命令行模式会打印统计与错误详情。

关键依赖：
- Python 3.10+
- openpyxl

最小使用示例：
    python3 parse_students_xlsx.py input/students.xlsx
"""
from __future__ import annotations
import sys
from pathlib import Path
from typing import Any

import openpyxl

# 29 列字段定义（顺序即列顺序）
SCHEMA: list[dict] = [
    # A. 标识与决策者
    {"name": "学生姓名", "type": "str", "required": True},
    {"name": "学生代号", "type": "str", "required": False},
    {"name": "主要决策者", "type": "enum", "required": True,
     "values": ["学生本人", "家长", "共同"]},
    # B. 基本学业信息
    {"name": "省份", "type": "str", "required": True},
    {"name": "户籍", "type": "str", "required": True},
    {"name": "高考年份", "type": "int", "required": True, "range": (2024, 2030)},
    {"name": "科类_选科", "type": "str", "required": True,
     "note": "老高考填 文科/理科；新高考填 物化生/物化地/物生地/史政地/史地生 等"},
    {"name": "总分", "type": "int", "required": True, "range": (0, 900)},
    {"name": "全省位次", "type": "int", "required": True, "range": (0, 1500000),
     "note": "0 表示未知，由 skill 用一分一段表反推"},
    {"name": "语文分", "type": "int", "required": True, "range": (0, 150)},
    {"name": "数学分", "type": "int", "required": True, "range": (0, 150)},
    {"name": "外语分", "type": "int", "required": True, "range": (0, 150)},
    {"name": "外语语种", "type": "enum", "required": False,
     "values": ["英语", "日语", "俄语", "德语", "法语", "西语"], "default": "英语"},
    # C. 兴趣与偏好
    {"name": "兴趣方向", "type": "str", "required": True},
    {"name": "厌恶清单", "type": "str", "required": False},
    {"name": "梦想校", "type": "str", "required": False},
    {"name": "性格倾向", "type": "enum", "required": True,
     "values": ["内向独立", "内向协作", "外向独立", "外向协作"]},
    {"name": "地域偏好", "type": "str", "required": True},
    {"name": "排斥地域", "type": "str", "required": False},
    {"name": "优先级排序", "type": "str", "required": True,
     "note": "形如 学校>专业>城市"},
    # D. 约束与意向
    {"name": "学费上限_万每年", "type": "float", "required": True,
     "note": "0 表示无限制"},
    {"name": "接受民办_独立学院", "type": "bool", "required": True},
    {"name": "接受中外合办", "type": "bool", "required": True},
    {"name": "服从专业调剂", "type": "bool", "required": True},
    {"name": "读研意向", "type": "enum", "required": True,
     "values": ["直接就业", "国内读研", "出国", "不确定"]},
    {"name": "家庭月收入_万", "type": "str", "required": True,
     "note": "区间 如 1-2 / 2-5 / 5+"},
    # E. 体检与特殊招生
    {"name": "体检受限项", "type": "str", "required": True,
     "note": "无 或如 色弱;视力<4.8（多值用分号）"},
    {"name": "特殊招生通道", "type": "str", "required": True,
     "note": "普通批 或如 强基:数学;综评:港中深"},
    {"name": "备注", "type": "str", "required": False},
]


def _to_bool(v: Any) -> bool | None:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("是", "true", "1", "y", "yes"):
        return True
    if s in ("否", "false", "0", "n", "no"):
        return False
    raise ValueError(f"无法识别的布尔值: {v!r}")


def _coerce(value: Any, col: dict) -> Any:
    """按 schema 转换单个字段值，错误抛 ValueError。"""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        if col["required"]:
            raise ValueError(f"必填字段 {col['name']} 为空")
        return col.get("default")

    t = col["type"]
    try:
        if t == "str":
            v = str(value).strip()
        elif t == "int":
            v = int(float(value))  # 兼容 670.0 这类
            lo, hi = col.get("range", (None, None))
            if lo is not None and (v < lo or v > hi):
                raise ValueError(f"{col['name']}={v} 超出范围 [{lo},{hi}]")
        elif t == "float":
            v = float(value)
        elif t == "bool":
            v = _to_bool(value)
        elif t == "enum":
            s = str(value).strip()
            if s not in col["values"]:
                raise ValueError(
                    f"{col['name']}={s!r} 不在枚举 {col['values']} 内"
                )
            v = s
        else:
            raise ValueError(f"未知 type {t}")
    except (ValueError, TypeError) as e:
        raise ValueError(f"{col['name']}={value!r}: {e}") from e
    return v


def parse_xlsx(path: str | Path) -> tuple[list[dict], list[dict]]:
    """读取 xlsx，返回 (rows, errors)。

    rows: 每行一个 dict，字段名为 SCHEMA["name"]
    errors: 行错误清单 [{"row": 2, "errors": ["..."]}]
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    header = [c.value for c in ws[1]]
    schema_names = [c["name"] for c in SCHEMA]

    # 校验表头
    missing = set(schema_names) - set(header or [])
    if missing:
        raise ValueError(
            f"Excel 表头缺失字段: {sorted(missing)}。"
            f"必须包含 {schema_names}"
        )

    # 列名 → 列索引
    name2idx = {n: i for i, n in enumerate(header)}

    rows, errors = [], []
    for r_idx, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True), start=2
    ):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        rec: dict = {"_row": r_idx}
        row_errs: list[str] = []
        for col in SCHEMA:
            raw = row[name2idx[col["name"]]] if col["name"] in name2idx else None
            try:
                rec[col["name"]] = _coerce(raw, col)
            except ValueError as e:
                row_errs.append(str(e))
                rec[col["name"]] = None
        if row_errs:
            errors.append({"row": r_idx, "name": rec.get("学生姓名"), "errors": row_errs})
        rows.append(rec)
    return rows, errors


def print_summary(rows: list[dict], errors: list[dict]) -> None:
    """打印解析摘要。"""
    print(f"成功解析: {len(rows)} 行学生数据")
    print(f"含错误的行: {len(errors)}")
    for e in errors:
        print(f"  L{e['row']} {e['name']}: {'; '.join(e['errors'])}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: parse_students_xlsx.py <path/to/students.xlsx>")
        sys.exit(1)
    rows, errors = parse_xlsx(sys.argv[1])
    print_summary(rows, errors)
    if rows:
        import json
        print("\n首行示例:")
        sample = {k: v for k, v in rows[0].items() if k != "_row"}
        print(json.dumps(sample, ensure_ascii=False, indent=2))
